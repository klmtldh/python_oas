import openpyxl
import pandas as pd
class OasXlsx(object):
    def __init__(self,
                 read_excel_path,
                 read_excel_sheet,
                 save_excel_path,
                 save_excel_sheet):
        self.read_excel_name = read_excel_path
        self.read_excel_sheet = read_excel_sheet
        self.save_excel_name = save_excel_path
        self.save_excel_sheet = save_excel_sheet

    def read_excel_file_to_dataframe(self):
        if self.read_path.endswith('.xlsx'):
            excel_file = pd.ExcelFile(self.read_excel_path, engine='openpyxl')
        elif self.read_path.endswith('.xls'):
            excel_file = pd.ExcelFile(self.read_excel_path)
        else:
            return None
        sheet = excel_file(self.save_excel_sheet)
        return sheet

    # 读取 excel文件
    def read_excel_file_to_list(self, path: str, sheet: str):
        # 实例化一个workbook对象
        workbook = openpyxl.load_workbook(self.read_excel_path)
        # 获取excel文件内的那一个sheet
        data = workbook[self.read_excel_sheet]
        # 定义个要输出的总数组
        return_data = []
        # data.rows 为表格内的每一行数据
        # 循环获取表格内的每一行数据
        for index, row in enumerate(data.rows):
            # 定义一个空的数组用来存放每一行数据单元格的数据
            return_row = []
            for col_index, col_value in enumerate(row):
                # 获取单元格数据 追加到return_row
                return_row.append(col_value.value)
            return_data.append(return_row)
        # 把遍历出来得每一行数据数据return_row 追加到总数组 return_data中 然后输出

        return return_data

    def make_excel_file(self):
        # 实例化
        self.wb = openpyxl.Workbook()
        # 激活 worksheet
        ws = wb.active

    def open_excel_file(self):
        self.wb = openpyxl.load_workbook('文件名称.xlsx')



